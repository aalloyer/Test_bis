#%% PRESTART
    
import numpy as np
import pandas as pd
from scipy.interpolate import interp1d
import os
import netCDF4 as nc
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st
from io import BytesIO
import matplotlib.pyplot as plt
import datetime
import warnings

warnings.filterwarnings("ignore", message="numpy.dtype size changed")
warnings.filterwarnings("ignore", message="numpy.ufunc size changed")


# functions
@st.cache_data
def CDFt_ds(GF, GH, SH):
    current_time = datetime.datetime.now()
    st.write(f"CDFt_ds en exécution à {current_time}")
    cdf_GF = np.sort(GF)
    u_GF = np.searchsorted(cdf_GF, GF, side='right') / len(cdf_GF)
    x_GH = np.quantile(GH, u_GF)
    cdf_SH = np.sort(SH)
    u_SH = np.searchsorted(cdf_SH, x_GH, side='right') / len(cdf_SH)

    unique_u_SH = np.unique(u_SH)
    mean_GF_per_unique_u_SH = [np.mean(GF[u_SH == u]) for u in unique_u_SH]

    cdf_SF_inverse = interp1d(unique_u_SH, mean_GF_per_unique_u_SH, bounds_error=False, fill_value="extrapolate")

    interpolated_values = cdf_SF_inverse(u_GF)
    interpolated_values_full = pd.Series(interpolated_values).interpolate(method='linear').to_numpy()
    return interpolated_values_full

@st.cache_data
def Qmap_ds(GF, GH, SH):
    current_time = datetime.datetime.now()
    st.write(f"Qmap_ds en exécution à {current_time}")
    cdf_GH = np.sort(GH)
    SF = np.quantile(SH, np.searchsorted(cdf_GH, GF, side='right') / len(cdf_GH))
    return SF

@st.cache_data
def nearest_point_cmip(lon_, lat_, path_nc): 
    current_time = datetime.datetime.now()
    st.write(f"nearest_point_cmip en exécution à {current_time}")
    files = os.listdir(path_nc)
    nc_input = nc.Dataset(f"{path_nc}{files[0]}")
    lon = nc_input.variables["lon"][:]
    lat = nc_input.variables["lat"][:]
    time_var = nc_input.variables["time"]
    time = nc.num2date(time_var[:], 
                       units= time_var.units,
                       calendar = time_var.calendar )
    time = time.astype(str).tolist()
    time = pd.to_datetime(time)
    nearest_lon = np.argmin(np.abs(lon - lon_))
    nearest_lat = np.argmin(np.abs(lat - lat_))
    st.write(lon[nearest_lon])
    st.write(lat[nearest_lat])
    temperature = nc_input.variables["tas"][:,nearest_lat,nearest_lon]

    return pd.Series(temperature, index=time)
  
@st.cache_data
def plot_annualcycle(local_annual_cycle, cmip_annual_cycle, mean_or_max, period, model_or_mean):
    current_time = datetime.datetime.now()
    st.write(f"plot_annualcycle en exécution à {current_time}")
    fig, ax = plt.subplots()
    ax.plot(local_annual_cycle.index, local_annual_cycle.values, label='Local')
    ax.plot(cmip_annual_cycle.index, cmip_annual_cycle.values, label='Global')
    ax.legend() 
    ax.set_title(f"{mean_or_max} - {model_or_mean} - {period}")
    plt.grid(True)
    plot_obj = (fig, ax)
    plt.show()
    return plot_obj

@st.cache_data
def excel_processing(lon_, lat_, case_study) :
    current_time = datetime.datetime.now()
    st.write(f"excel_processing en exécution à {current_time}")
    case_study_mast_df = pd.DataFrame({
        'Date': case_study['TimeStamp'],
        'temperature': case_study['T° 4m LT dowscaled'],
        'windspeed': case_study['WS 120m LT reconst MCP']
    })
    case_study_mast_df = case_study_mast_df.iloc[1:] 
    case_study_mast_df['temperature'] = pd.to_numeric(case_study_mast_df['temperature'], errors='coerce')
    case_study_mast_df['windspeed'] = pd.to_numeric(case_study_mast_df['windspeed'], errors='coerce')
    case_study_mast_hourly_xts = pd.Series(case_study_mast_df['temperature'].values,
                                           index=pd.to_datetime(case_study_mast_df['Date']))
    case_study_mast_hourly_windspeed_xts = pd.Series(case_study_mast_df['windspeed'].values,
                                           index=pd.to_datetime(case_study_mast_df['Date']))


    
    start_year_brut = int(case_study_mast_hourly_xts.index[0].strftime('%Y'))
    hours_in_start_year_brut = pd.date_range(start=f"{start_year_brut}-01-01", end=f"{start_year_brut+1}-01-01", freq='H')
    
    end_year_brut = int(case_study_mast_hourly_xts.index[-1].strftime('%Y'))
    hours_in_end_year_brut = pd.date_range(start=f"{end_year_brut}-01-01", end=f"{end_year_brut+1}-01-01", freq='H')
    
    
    if case_study_mast_hourly_xts.index.isin(hours_in_start_year_brut).all():
        st.write(f"L'année {start_year_brut} est pleine. La TS commence en janvier {start_year_brut}.")
        start_bool = False
    else:
        st.write(f"L'année {start_year_brut} n'est pas pleine. La TS commence en janvier {start_year_brut+1}.")
        start_bool = True
    
    if case_study_mast_hourly_xts.index.isin(hours_in_end_year_brut).all():
        st.write(f"L'année {end_year_brut} est pleine. La TS termine en décembre {end_year_brut}.")
        end_bool = False
    else:
        st.write(f"L'année {end_year_brut} n'est pas pleine. La TS termine en décembre {end_year_brut-1}.")
        end_bool = True
        
    if start_bool :
        case_study_mast_hourly_xts = case_study_mast_hourly_xts[f"{start_year_brut+1}-01-01":]
        case_study_mast_hourly_windspeed_xts = case_study_mast_hourly_windspeed_xts[f"{start_year_brut+1}-01-01":]
    if end_bool :
        case_study_mast_hourly_xts = case_study_mast_hourly_xts[:f"{end_year_brut-1}-12-31"]
        case_study_mast_hourly_windspeed_xts = case_study_mast_hourly_windspeed_xts[:f"{end_year_brut-1}-12-31"]
    
    # s'arrête à 23h du dernier jour de l'anée
    start_year = int(case_study_mast_hourly_xts.index[0].strftime('%Y'))
    end_year = int(case_study_mast_hourly_xts.index[-1].strftime('%Y'))
    
    # Création des TS monthly pour DS, et annual pour linear model
    
    case_study_mast_monthly_xts = case_study_mast_hourly_xts.resample('M').mean()
    # case_study_mast_monthly_xts = case_study_mast_monthly_xts[case_study_mast_monthly_xts.index <= "2022-12-31 23:59:59"]
    case_study_mast_annual_xts = case_study_mast_hourly_xts.resample('A').mean()

    mean_temperature_historical = np.mean(case_study_mast_annual_xts.values)
    mean_temperature_historical = round(mean_temperature_historical, 2)
    st.write(f"Mean temperature on past period (from {start_year}-01 to {end_year}-12) : {mean_temperature_historical}°C")

    path_data = os.path.abspath(os.path.join(os.path.dirname(__file__), 'data/Near_surface_air_temperature')) 
    #fichiers_data = os.listdir(path_data)
    #st.write(fichiers_data)
    
    model_list = ['bcc_csm2_mr',
                  'cnrm_esm2_1',
                  'fgoals_g3',
                  'gfdl_esm4',
                  'ipsl_cm6a_lr',
                  'mri_esm2_0']
    output_tot =[]
    temperature_hist_xts_tot = []
    temperature_proj_xts_tot = []
    regression_df_tot =[]

    duration = case_study_mast_hourly_xts.index[-1] - case_study_mast_hourly_xts.index[0] + pd.Timedelta(hours=1)
    period_proj_1 = case_study_mast_hourly_xts.index + duration
    period_proj_2 = case_study_mast_hourly_xts.index + 2*duration
    period_proj_3 = case_study_mast_hourly_xts.index + 3*duration

    years_hist = case_study_mast_annual_xts.index.year
    X_hist =sm.add_constant(years_hist)
    regression_hist = sm.OLS(case_study_mast_annual_xts, X_hist).fit()
    slope_hist = regression_hist.params[1]
    r2_hist = regression_hist.rsquared
    regression_hist = [slope_hist,
                       r2_hist]

    del duration, X_hist
    index_count = 0
    flat_data = case_study_mast_hourly_xts.values.copy()
    
    for year in years_hist :
        hours_in_year =  len(pd.date_range(start=f"{year}-01-01 00:00:00", end=f"{year}-12-31 23:00:00", freq='H'))
        for hour in range(0, hours_in_year):
            flat_data[index_count] = case_study_mast_hourly_xts.values[index_count] - slope_hist*(year + hour/hours_in_year - start_year)
            index_count =  index_count + 1
    del hour, year, hours_in_year
    
    for model in model_list :
            
            #CMIP DATA TREATMENT - HISTORICAL 
            path_cmip_hist = f"{path_data}/historical/{model}_historical/"   
            temperature_hist_xts = nearest_point_cmip(lon_=lon_, lat_=lat_,path_nc = path_cmip_hist)
            temperature_hist_xts = temperature_hist_xts - 273.15
            temperature_hist_xts = temperature_hist_xts[
                case_study_mast_hourly_xts.index[0]: case_study_mast_hourly_xts.index[-1]
                ]
    
            
            #CMIP DATA TREATMENT - PROJECTION 
            path_cmip_proj = f"{path_data}/ssp3_7_0/{model}_ssp3_7_0/"
            temperature_proj_xts = nearest_point_cmip(lon_=lon_, lat_=lat_, path_nc = path_cmip_proj)
            temperature_proj_xts = temperature_proj_xts - 273.15
            temperature_proj_xts_tot.append(temperature_proj_xts)


        
            # to complete historical time series
            if end_year > 2014:
                temperature_hist2_xts = temperature_proj_xts[
                case_study_mast_hourly_xts.index[0]: case_study_mast_hourly_xts.index[-1]
                ]
                new_index = np.concatenate((temperature_hist_xts.index, temperature_hist2_xts.index))
                new_values = np.concatenate((temperature_hist_xts.values, temperature_hist2_xts.values))
                temperature_hist_xts = pd.Series(new_values, index = new_index)
            
            temperature_hist_xts_tot.append(temperature_hist_xts) 

            
            # ATTENTION AUX HEURES - ok car données horaires
            #to take projected proj time series
            temperature_proj_1_xts = temperature_proj_xts[period_proj_1[0]:period_proj_1[-1]]
            temperature_proj_2_xts = temperature_proj_xts[period_proj_2[0]:period_proj_2[-1]]
            temperature_proj_3_xts = temperature_proj_xts[period_proj_3[0]:period_proj_3[-1]]
            
            temperature_proj_xts_list = [temperature_proj_1_xts, 
                                         temperature_proj_2_xts,
                                         temperature_proj_3_xts]
            
            # LINEAR MODEL
            
            # Future 
            years_list = []
            regression_list = []
            
            for temperature_proj_i_xts in temperature_proj_xts_list :
                temperature_DS_proj_i_values = Qmap_ds(temperature_proj_i_xts,
                                                       temperature_hist_xts,
                                                       case_study_mast_monthly_xts)
                
                temperature_DS_proj_i_monthly_xts = pd.Series(temperature_DS_proj_i_values,
                                                      index = temperature_proj_i_xts.index)
                
                temperature_DS_proj_i_annual_xts = temperature_DS_proj_i_monthly_xts.resample('A').mean()
                years_i = temperature_DS_proj_i_annual_xts.index.year
                years_list.append(years_i)
                
                X_i = sm.add_constant(years_i)  # Adds a constant term to the predictor # pouequoi?
                regression_proj_i = sm.OLS(temperature_DS_proj_i_annual_xts, X_i).fit()
                
                slope_proj_i = regression_proj_i.params[1]
                #intercept_proj_1 = regression_proj_1.params[0]
                r2_proj_i = regression_proj_i.rsquared
                regression_data_i = [slope_proj_i,
                                     r2_proj_i]
                regression_list.append(regression_data_i)
            
            regression_df = pd.DataFrame(regression_list, columns=['slope', 'r2'], index=['period_proj_1', 'period_proj_2', 'period_proj_3'])
            
            
            # hours_in_end_year = len(pd.date_range(start=f"{end_year}-01-01 00:00:00", end=f"{end_year}-12-31 23:00:00", freq='H'))
            offset_hist = slope_hist * (end_year - start_year + 1) # pas un bon offset # en relatif
            offset_1 = regression_df.iloc[0]['slope'] * (end_year - start_year + 1) # même durée donc ok
            offset_2 = regression_df.iloc[1]['slope'] * (end_year - start_year + 1) # même durée donc ok
            
            proj_data_1 = flat_data + offset_hist
            proj_data_2 = flat_data + offset_hist + offset_1
            proj_data_3 = flat_data + offset_hist + offset_1 + offset_2
            
            proj_data_df = [proj_data_1.copy(),
                            proj_data_2.copy(),
                            proj_data_3.copy()]
    
            i = 0
            for years_i in years_list:
                index_count = 0
                slope_i = regression_df.iloc[i]['slope']
                
                for year in years_i :
                    hours_in_year = len(pd.date_range(start=f"{year}-01-01 00:00:00", end=f"{year}-12-31 23:00:00", freq='H'))
                    for hour in range(0, hours_in_year):
                        proj_data_df[i][index_count] = proj_data_df[i][index_count] + slope_i*(year+hour/hours_in_year - years_i[0])
                        index_count =  index_count + 1
                i = i + 1
                    
            case_study_proj_1_xts = pd.Series(proj_data_df[0],
                                               index = period_proj_1)
            case_study_proj_2_xts = pd.Series(proj_data_df[1],
                                               index = period_proj_2)
            case_study_proj_3_xts = pd.Series(proj_data_df[2],
                                               index = period_proj_3)
            
            output_tot.append([case_study_proj_1_xts, 
                               case_study_proj_2_xts, 
                               case_study_proj_3_xts])
    
    output_list = [output_tot,
                   temperature_hist_xts_tot,
                   temperature_proj_xts_tot,
                   regression_df_tot,
                   case_study_mast_hourly_windspeed_xts,
                   case_study_mast_hourly_xts]
    return output_list

@st.cache_data
def excel_downloading(output_tot, lon_, lat_, implementation_date, lifetime) :
    current_time = datetime.datetime.now()
    st.write(f"excel_downloading en exécution à {current_time}")
    output_df = pd.DataFrame(output_tot, index = model_list )
    output_all_list = []
    
    for i in range(0, len(model_list)):
        values = np.concatenate([output_df.iloc[i][0].values,
                                 output_df.iloc[i][1].values,
                                 output_df.iloc[i][2].values])
        index = np.concatenate([output_df.iloc[i][0].index,
                                output_df.iloc[i][1].index,
                                output_df.iloc[i][2].index])
        output_all_list.append(pd.Series(values, index = index))
        
    model_list.append('Mean')
    mean_values = np.mean([ts.values for ts in output_all_list], axis=0)
    mean_temperature_series = pd.Series(mean_values, index =output_all_list[0].index)
    output_all_list.append(mean_temperature_series)
    
    windspeed_tot = np.concatenate([case_study_mast_hourly_windspeed_xts.values,
                                   case_study_mast_hourly_windspeed_xts.values,
                                   case_study_mast_hourly_windspeed_xts.values])
    windspeed_tot_series = pd.Series(windspeed_tot, index = output_all_list[0].index)
    
    # Traitement de la série temporelle 
    implementation_month, implementation_year = implementation_date.split('/')
    final_year = int(implementation_year) + lifetime
    windfarm_start = f"{implementation_year}-{implementation_month}-01"
    windfarm_end = f"{str(final_year)}-{implementation_month}-01"
    
    if pd.to_datetime(windfarm_end) <= output_all_list[0].index[-1]:
        mean_temperature_extract_series = mean_temperature_series[windfarm_start:windfarm_end]
        windspeed_extract_series = windspeed_tot_series[windfarm_start:windfarm_end]
        
        output_extract_df = pd.DataFrame({
            "Timestamp": mean_temperature_extract_series.index,
            "Windspeed [m/s]": windspeed_extract_series.values, 
            "Temperature [deg C]": mean_temperature_extract_series.values
        }) 
    
        # extraction température moyenne et la slope
        mean_temperature_extract_value = np.mean(mean_temperature_extract_series.values)
        mean_temperature_extract_value = round(mean_temperature_extract_value, 2)
        st.write(f"Mean temperature on selected period (from {implementation_year}-{implementation_month} to {str(final_year)}-{implementation_month}) : {mean_temperature_extract_value}°C")
        
        if int(implementation_month) != 1 :
            for_slope_xts = mean_temperature_extract_series[f"{int(implementation_year)+1}-01-01":
                                                   f"{final_year-1}-12-31"]
            
            for_slope_annual_xts = for_slope_xts.resample('A').mean()
            for_slope_year = for_slope_annual_xts.index.year
            X_i = sm.add_constant(for_slope_year)  
            regression_extract= sm.OLS(for_slope_annual_xts, X_i).fit()
            slope_extract = regression_extract.params[1]
        else :
            for_slope_annual_xts = mean_temperature_extract_series.resample('A').mean()
            for_slope_year = for_slope_annual_xts.index.year
            X_i = sm.add_constant(for_slope_year)  
            regression_extract= sm.OLS(for_slope_annual_xts, X_i).fit()
            slope_extract = regression_extract.params[1]
            slope_extract = round(slope_extract, 2)
            st.write(f"Indicative annual average temperature increase on selected period (from {implementation_year}-{implementation_month} to {str(final_year)}-{implementation_month}): {slope_extract}°C/year")
        
    else :
        print("L'année de fin de vie dépasse la période prédite.")
        output_extract = 0
    
    output_list = [output_extract_df, mean_temperature_series, windfarm_start, windfarm_end, output_all_list] 
    return output_list
    
@st.cache_data
def validation_process(case_study_mast_hourly_xts, mean_temperature_series, temperature_proj_xts_tot, windfarm_start, windfarm_end, lon_, lat_, implementaton_date, lifetime) :
    current_time = datetime.datetime.now()
    st.write(f"validation_process en exécution à {current_time}")
    #%% VALIDATION PRESTART - observed data
    
    # TS_DS : on moyenne, prend le max mensuellement
    case_study_mast_monthly_mean_xts = case_study_mast_hourly_xts.resample('M').mean()
    case_study_mast_monthly_max_xts = case_study_mast_hourly_xts.resample('M').max()
    
    output_monthly_mean_xts = mean_temperature_series.resample('M').mean()
    output_monthly_max_xts = mean_temperature_series.resample('M').max()
    
    # regrouper en moyenne cycle annuel
    
    #%% VALIDATION PRSTART - climate data (=undownscaled)
    
    #MEAN
    
    # projection
    mean_brut_proj_values  = np.mean([ts.values for ts in temperature_proj_xts_tot], axis=0)
    mean_brut_proj_xts = pd.Series(mean_brut_proj_values, index =temperature_proj_xts_tot[0].index)
    # temperature_proj_xts_tot.append(mean_brut_proj_xts) 
    del mean_brut_proj_values
    mean_brut_proj_xts = mean_brut_proj_xts[windfarm_start:windfarm_end] # relevant period
    
    
    #historical
    mean_brut_hist_values =  np.mean([ts.values for ts in temperature_hist_xts_tot], axis=0)
    mean_brut_hist_xts = pd.Series(mean_brut_hist_values, index =temperature_hist_xts_tot[0].index)
    # temperature_hist_xts_tot.append(mean_brut_hist_xts) 
    del mean_brut_hist_values
    
    #MEAN VALIDATION - bien vérifier que cest par mois / attention la relevant period est seulement à faire pour projected
    
    # LOCAL DATA
    
    # HISTORICAL 
    #case_study_mast_monthly_mean_xts
    
    local_historical_mean_df = pd.DataFrame({'month':case_study_mast_monthly_mean_xts.index.month,
                                             'year': case_study_mast_monthly_mean_xts.index.year,
                                             'mean_temperature': case_study_mast_monthly_mean_xts.values
                                             })
    local_historical_mean_annual_cycle_mean = local_historical_mean_df.groupby('month')['mean_temperature'].mean()
    del local_historical_mean_df
    
    
    # PROJECTED
    
    #output_monthly_mean_xts : MEAN
    
    local_projected_mean_df = pd.DataFrame({'month':output_monthly_mean_xts.index.month,
                                             'year': output_monthly_mean_xts.index.year,
                                             'mean_temperature': output_monthly_mean_xts.values
                                             })
    local_projected_mean_annual_cycle_mean = pd.DataFrame(local_projected_mean_df.groupby('month')['mean_temperature'].mean())
    del local_projected_mean_df
    
    
    #for each model : 
    output_all_list_extract_mean = [] # proj for each model 
    output_all_list_extract_max =[]
    for series in output_all_list :
        new_series = series[windfarm_start:windfarm_end]
        new_series_mean = new_series.resample('M').mean()
        new_series_max = new_series.resample('M').max()
        output_all_list_extract_mean.append(new_series_mean)
        output_all_list_extract_max.append(new_series_max)
        del series, new_series
    
    local_projected_model_annual_cycle_mean = []
    for series in output_all_list_extract_mean :
        df =  pd.DataFrame({'month':series.index.month,
                                                'year': series.index.year,
                                                'mean_temperature': series.values
                                                })
        df_annual_cycle = pd.DataFrame(df.groupby('month')['mean_temperature'].mean())
        local_projected_model_annual_cycle_mean.append(df_annual_cycle)
    
    
    
    # CLIMATE DATA 
    
    # HISTORICAL
     
    # mean_brut_hist_xts: MEAN
    
    cmip_historical_mean_df = pd.DataFrame({'month':mean_brut_hist_xts.index.month,
                                             'year': mean_brut_hist_xts.index.year,
                                             'mean_temperature': mean_brut_hist_xts.values
                                             })
    cmip_historical_mean_annual_cycle_mean = cmip_historical_mean_df.groupby('month')['mean_temperature'].mean()
    del cmip_historical_mean_df
    
    
    # for each model :
    # temperature_hist_xts_tot_extract = []
    # for series in temperature_hist_xts_tot : 
    #     new_series = series[windfarm_start:windfarm_end]
    #     temperature_hist_xts_tot_extract.append(new_series)
    #     del series, new_series
    
    cmip_historical_model_annual_cycle_mean = []
    for series in temperature_hist_xts_tot :
        df =  pd.DataFrame({'month':series.index.month,
                                                'year': series.index.year,
                                                'mean_temperature': series.values
                                                })
        df_annual_cycle = pd.DataFrame(df.groupby('month')['mean_temperature'].mean())
        cmip_historical_model_annual_cycle_mean.append(df_annual_cycle)
    
    
    # PROJECTED
    
    # mean_brut_proj_xts : MEAN 
    
    cmip_proj_mean_df = pd.DataFrame({'month':mean_brut_proj_xts.index.month,
                                             'year': mean_brut_proj_xts.index.year,
                                             'mean_temperature': mean_brut_proj_xts.values
                                             })
    cmip_proj_mean_annual_cycle_mean = cmip_proj_mean_df.groupby('month')['mean_temperature'].mean()
    del cmip_proj_mean_df
    
    # for each model :
    temperature_proj_xts_tot_extract = []
    for series in temperature_proj_xts_tot : 
        new_series = series[windfarm_start:windfarm_end]
        temperature_proj_xts_tot_extract.append(new_series)
        del series, new_series
    
    cmip_projected_model_annual_cycle_mean = []
    for series in temperature_proj_xts_tot_extract :
        df =  pd.DataFrame({'month':series.index.month,
                                                'year': series.index.year,
                                                'mean_temperature': series.values
                                                })
        df_annual_cycle = pd.DataFrame(df.groupby('month')['mean_temperature'].mean())
        cmip_projected_model_annual_cycle_mean.append(df_annual_cycle)
        
    
    #%% PLOT
    
    dico_local = {
        'mean_historical_mean' :local_historical_mean_annual_cycle_mean,
        'mean_historical_model' : local_historical_mean_annual_cycle_mean,
        'mean_projected_mean': local_projected_mean_annual_cycle_mean,
        'mean_projected_model':local_projected_model_annual_cycle_mean}
    
    dico_global = {
        'mean_historical_mean' : cmip_historical_mean_annual_cycle_mean,
        'mean_historical_model' : cmip_historical_model_annual_cycle_mean,
        'mean_projected_mean': cmip_proj_mean_annual_cycle_mean,
        'mean_projected_model':cmip_projected_model_annual_cycle_mean}

    dico_tot = [dico_local, dico_global, mean_brut_hist_xts, mean_brut_proj_xts]
    return dico_tot 
    
#%%INPUTS

st.write("ligne 515")

# geographical data input
st.sidebar.title("INPUTS")

# afficher hypothèses sur longitude et latitude
lon_site = st.sidebar.number_input("Longitude (from 0° to 360°) : ", step=0.1)
lat_site = st.sidebar.number_input("Latitude (from -90° [south] to 90° [north])", step=0.1) # intervalle ? 
implementation_date = st.sidebar.text_input("Commission date (MM/YYYY)")
lifetime = st.sidebar.number_input("Wind farm lifetime (in years)", step=1)  
model_list = ['bcc_csm2_mr',
              'cnrm_esm2_1',
              'fgoals_g3',
              'gfdl_esm4',
              'ipsl_cm6a_lr',
              'mri_esm2_0']

st.write("ligne 532")

#%% Interface
telecharge = False
if 'uploaded_data' not in st.session_state:
    st.session_state['uploaded_data'] = None

uploaded_file = st.file_uploader("Choose a file")

if uploaded_file is not None:
    st.session_state['uploaded_data'] = pd.read_excel(uploaded_file,
                            sheet_name="Reconst", 
                            skiprows=3)
    case_study_mast = st.session_state['uploaded_data']
    output_list = excel_processing(lon_site, lat_site, case_study_mast)
    output_tot = output_list[0]
    temperature_hist_xts_tot = output_list[1]
    temperature_proj_xts_tot = output_list[2]
    regression_df_tot = output_list[3]
    case_study_mast_hourly_windspeed_xts = output_list[4]
    case_study_mast_hourly_xts = output_list[5]

    output_excel_downloading =  excel_downloading(output_tot, lon_site, lat_site, implementation_date, lifetime)
    output_extract_df = output_excel_downloading[0]
    mean_temperature_series  = output_excel_downloading[1]
    windfarm_start = output_excel_downloading[2]
    windfarm_end = output_excel_downloading[3]
    output_all_list = output_excel_downloading[4]

    st.write("ligne 561")
    
    if not output_extract_df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = "Temperature Projection - extract period"
        for r in dataframe_to_rows(output_extract_df, index=False, header=True):
                ws.append(r)
        wb.save("Temperature Projection - extract period.xlsx")
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        st.download_button(
            label="Télécharger le fichier Excel",
            data=output,
            file_name="T&WS-TS.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.write("ligne 580")
    # Validation
    dico_tot = validation_process(case_study_mast_hourly_xts,mean_temperature_series, temperature_proj_xts_tot,windfarm_start, windfarm_end, lon_site, lat_site, implementation_date, lifetime)
    dico_local = dico_tot[0]
    dico_global = dico_tot[1]
    mean_brut_hist_xts = dico_tot[2]
    mean_brut_proj_xts = dico_tot[3]
    
    model_list_max = ['cnrm_esm2_1',
                  'fgoals_g3',
                  'gfdl_esm4',
                  'ipsl_cm6a_lr',
                  'mri_esm2_0']
    
    model_list = ['bcc_csm2_mr',
                  'cnrm_esm2_1',
                  'fgoals_g3',
                  'gfdl_esm4',
                  'ipsl_cm6a_lr',
                  'mri_esm2_0']
    
    mean_or_max = st.selectbox('Mean ou Max ?', ['','mean', 'max'])
    historical_or_projected = st.selectbox('Historical ou Projected ?', ['','historical', 'projected'])
    mean_or_model = st.selectbox('Mean ou Model ?', ['','mean', 'bcc_csm2_mr', 'cnrm_esm2_1','fgoals_g3','gfdl_esm4','ipsl_cm6a_lr','mri_esm2_0'])
    p = 0
    
    # attention si = "max" et que = au model qu'il n'y a pas
    if (mean_or_max == "max") and (mean_or_model == "bcc_csm2_mr") :
      st.write("Le modèle bcc_csm2_mr ne fournit pas de données de température maximale.")
    elif mean_or_model == "mean":
        object_name = f"{mean_or_max}_{historical_or_projected}_{mean_or_model}"
        local = dico_local[object_name]
        cmip = dico_global[object_name]
        p = plot_annualcycle(local, cmip, mean_or_max, historical_or_projected, mean_or_model)
    else :
        if mean_or_max == "max":
            index = model_list_max.index(mean_or_model)
        else :
            index = model_list.index(mean_or_model)
        object_name = f"{mean_or_max}_{historical_or_projected}_model"
        local = dico_local[object_name] 
        cmip = dico_global[object_name] # pq c'est qu'une liste
        if len(local) == 12 :
            p = plot_annualcycle(local, cmip[index], mean_or_max, historical_or_projected, mean_or_model)
        else :
            p = plot_annualcycle(local[index], cmip[index], mean_or_max, historical_or_projected, mean_or_model)

    if p !=0 :
        my_fig = p[0]
        my_ax = p[1]
        st.pyplot(my_fig)

    
    #%% CMIP TS DOWNLOADING  : SEPARER PROJ DE HIST 
    
    mean_cmip_tot = []
    
    mean_brut_values = np.concatenate((mean_brut_hist_xts.values, mean_brut_proj_xts.values))
    date = np.concatenate((mean_brut_hist_xts.index, mean_brut_proj_xts.index))
    cmip_mean_xts = pd.Series(mean_brut_values, index =date)
    mean_cmip_tot.append(cmip_mean_xts)



  
    for i in range(len(model_list)):
        temperature_proj_xts_tot[i] = temperature_proj_xts_tot[i][windfarm_start:windfarm_end]

    temperature_proj_xts_tot.append(mean_brut_proj_xts)
    temperature_hist_xts_tot.append(mean_brut_hist_xts)
    model_list = ['bcc_csm2_mr',
                  'cnrm_esm2_1',
                  'fgoals_g3',
                  'gfdl_esm4',
                  'ipsl_cm6a_lr',
                  'mri_esm2_0',
                  'mean']

    st.write("ligne 658")

    
    if st.checkbox('Climate models projected monthly mean temperature TS extraction (Downloading may take time and is not recommended.'):
      historical_checked = st.checkbox('Historical period')
      projected_checked = st.checkbox('Projected period')

      if historical_checked : 
        
        wb = Workbook()
        for i in range(len(model_list)):
            ws = wb.create_sheet(title=f"{model_list[i]}")
            data = {
                "Timestamp" : temperature_hist_xts_tot[i].index,
                "Temperature [Deg C]": temperature_hist_xts_tot[i].values
                }
            df = pd.DataFrame(data)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        st.download_button(label='Télécharger les données', 
                           data=output, 
                           file_name='cmip_mean_TS.xlsx', 
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

      if projected_checked : 
        wb = Workbook()
        for i in range(len(model_list)):
            ws = wb.create_sheet(title=f"{model_list[i]}")
            data = {
                "Timestamp" : temperature_proj_xts_tot[i].index,
                "Temperature [Deg C]": temperature_proj_xts_tot[i].values
                }
            df = pd.DataFrame(data)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        st.download_button(label='Télécharger les données', 
                           data=output, 
                           file_name='cmip_mean_TS.xlsx', 
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
